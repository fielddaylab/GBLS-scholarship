#!/usr/bin/env python3
"""
calculate_summary_metrics.py

Parses every coded article summary in 1_coded_gbls_corpus_articles/, calculates
reproducible metrics, and writes all artifacts to
2_calculated_metrics/gbls_corpus_metrics/.

Also regenerates the metrics_explorer_synthesized_data.js bundle.

Usage (from any directory):
    python3 /path/to/tools/calculate_summary_metrics.py

Dependencies: pandas, openpyxl
    pip install pandas openpyxl
"""

import re
import json
import math
import sys
import shutil
from pathlib import Path
from datetime import datetime, timezone
from collections import defaultdict, OrderedDict
from itertools import combinations

import pandas as pd

# ── Paths ───────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

SUMMARIES_DIR = PROJECT_ROOT / "1_coded_gbls_corpus_articles"
HUMAN_SOURCES_DIR = PROJECT_ROOT / "0_human_sources"
METRICS_DIR = PROJECT_ROOT / "2_calculated_metrics" / "gbls_corpus_metrics"
EXPLORER_DATA = PROJECT_ROOT / "2_calculated_metrics" / "metrics_explorer_synthesized_data.js"
EXPLORER_DATA_CODER = PROJECT_ROOT / "tools" / "gbls_lit_coder" / "public" / "2_outputs" / "metrics" / "metrics_explorer_synthesized_data.js"
EXPLORER_HTML = PROJECT_ROOT / "tools" / "metrics_explorer" / "index.html"
EXPLORER_JS = PROJECT_ROOT / "tools" / "metrics_explorer" / "metrics_explorer.js"

IDENTITY_COLS = ["article_id", "year", "decade", "citation", "summary_word_count"]


# ── Schema parsing ──────────────────────────────────────────────────────────

def find_schema_file(human_sources_dir: Path) -> Path:
    """Find the metadata schema / controlled lexicon file."""
    candidates = sorted(human_sources_dir.glob("*.md"))
    best = None
    best_score = 0
    for f in candidates:
        if f.name.startswith("."):
            continue
        text = f.read_text(encoding="utf-8", errors="replace")
        score = text.count("\n# ") + ("lexicon" in f.name.lower()) * 5 + ("schema" in f.name.lower()) * 5
        if score > best_score:
            best_score = score
            best = f
    if best is None:
        raise FileNotFoundError("No schema file found in 0_human_sources/")
    return best


def parse_schema(human_sources_dir: Path) -> tuple[list[str], dict[str, list[str]]]:
    """
    Parse the controlled lexicon file.
    Returns:
        field_names  – ordered list of controlled field names (as written in schema)
        field_values – dict mapping field_name -> list of allowed value strings
    """
    schema_file = find_schema_file(human_sources_dir)
    text = schema_file.read_text(encoding="utf-8", errors="replace")

    field_names: list[str] = []
    field_values: dict[str, list[str]] = {}
    current_field: str | None = None

    for line in text.split("\n"):
        # Top-level heading = field name (e.g. "# Source_Type" or "# Peer_Review:")
        m = re.match(r"^#\s+([A-Za-z_][A-Za-z0-9_]*)\s*:?\s*$", line)
        if m:
            current_field = m.group(1)
            if current_field not in field_names:
                field_names.append(current_field)
                field_values[current_field] = []
            continue

        if current_field is not None:
            # "- value_name: description" or "- value_name"
            m = re.match(r"^-\s+([a-zA-Z][a-zA-Z0-9_]*)", line)
            if m:
                val = m.group(1)
                if val not in field_values[current_field]:
                    field_values[current_field].append(val)

    return field_names, field_values


# ── Source file parsing ─────────────────────────────────────────────────────

def _collect_list_items(lines: list[str], start: int) -> tuple[list[str], int]:
    """
    Starting just after a 'Key:' line with no inline value, collect
    subsequent '- item' lines. Returns (items, next_line_index).
    Stops at blank lines followed by non-list content or at a new Key: line.
    """
    items: list[str] = []
    i = start
    while i < len(lines):
        line = lines[i]
        list_m = re.match(r"^\s*-\s+(.*)", line)
        if list_m:
            item = list_m.group(1).strip()
            # Skip YAML sub-key lines inside contribution blocks
            if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*\s*:", item):
                items.append(item)
            i += 1
        elif line.strip() == "":
            # Blank line – peek ahead
            peek = i + 1
            while peek < len(lines) and lines[peek].strip() == "":
                peek += 1
            if peek < len(lines) and re.match(r"^\s*-\s+", lines[peek]):
                i = peek  # skip blanks and continue collecting
            else:
                break  # blank line ends the list
        else:
            break  # non-list, non-blank line ends the list
    return items, i


def _parse_contributions(contrib_text: str) -> list[dict]:
    """
    Parse the YAML-like Contributions block into a list of dicts with
    keys: target_section, contribution_text.
    """
    contributions = []

    # Split into contribution blocks on lines that start a new list item
    # with 'Target_Section' key
    raw_blocks = re.split(r"\n(?=\s*-\s+Target_Section\s*:)", contrib_text)
    for block in raw_blocks:
        if not block.strip():
            continue
        # Remove the leading '- ' marker
        block = re.sub(r"^\s*-\s+", "", block, count=1)

        # Target_Section
        ts_m = re.search(r"^Target_Section\s*:\s*(.+)$", block, re.MULTILINE)
        target_section = ts_m.group(1).strip().strip('"') if ts_m else ""

        # Contribution_Text – may be folded (>) or inline
        ct_m = re.search(
            r"Contribution_Text\s*:\s*>?\s*\n((?:[ \t]+.+\n?)*)",
            block,
        )
        if ct_m:
            raw_ct = ct_m.group(1)
            # Collapse folded YAML: strip leading spaces, join lines with space
            contrib_text_val = re.sub(r"\n\s*", " ", raw_ct).strip()
        else:
            # Try inline value
            ct_inline = re.search(r"Contribution_Text\s*:\s*(.+)", block)
            contrib_text_val = ct_inline.group(1).strip() if ct_inline else ""

        # Skip placeholder "no contribution" entries
        if contrib_text_val.lower() in (
            "no_direct_addition_recommended",
            "no direct addition recommended",
            "none",
            "",
        ):
            continue

        contributions.append(
            {
                "target_section": target_section,
                "contribution_text": contrib_text_val,
            }
        )

    return contributions


def parse_summary_file(path: Path, schema_fields: list[str]) -> dict:
    """
    Parse a single coded summary .md file.
    Returns a record dict. Sets '_error' key on critical failure.
    """
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        return {"_error": str(exc), "_path": str(path), "filename": path.name}

    warnings: list[str] = []

    # ── Find section boundaries ────────────────────────────────────────────
    meta_m = re.search(r"^#{1,4}\s+Metadata\s*$", text, re.MULTILINE | re.IGNORECASE)
    summ_m = re.search(r"^#{1,4}\s+Summary\s*$", text, re.MULTILINE | re.IGNORECASE)

    if not meta_m:
        warnings.append("No Metadata heading found")

    if meta_m:
        citation_raw = text[: meta_m.start()].strip()
        meta_end = summ_m.start() if summ_m else len(text)
        meta_section = text[meta_m.end() : meta_end]
        summary_text = text[summ_m.end() :].strip() if summ_m else ""
    else:
        citation_raw = ""
        contrib_m = re.search(r"^Contributions\s*:", text, re.MULTILINE | re.IGNORECASE)
        meta_section = text[: contrib_m.start()] if contrib_m else text
        summary_text = ""

    # ── Clean citation ─────────────────────────────────────────────────────
    citation = re.sub(r"^#+\s*", "", citation_raw)
    citation = re.sub(r"^\d+\.\s*", "", citation).strip()

    # ── Split metadata from contributions ──────────────────────────────────
    contrib_start_m = re.search(
        r"^Contributions\s*:", meta_section, re.MULTILINE | re.IGNORECASE
    )
    if contrib_start_m:
        raw_meta = meta_section[: contrib_start_m.start()]
        raw_contrib = meta_section[contrib_start_m.end() :]
    else:
        raw_meta = meta_section
        raw_contrib = ""

    # ── Parse key-value metadata ───────────────────────────────────────────
    raw_fields: dict[str, str | list[str] | None] = {}
    lines = raw_meta.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i]
        kv_m = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)\s*:\s*(.*)", line)
        if kv_m:
            key = kv_m.group(1)
            value = kv_m.group(2).strip()
            if value:
                raw_fields[key] = value
                i += 1
            else:
                items, next_i = _collect_list_items(lines, i + 1)
                raw_fields[key] = items if items else None
                i = next_i
        else:
            i += 1

    # ── Parse contributions ────────────────────────────────────────────────
    contributions = _parse_contributions(raw_contrib) if raw_contrib.strip() else []

    # ── Identifiers ───────────────────────────────────────────────────────
    zotero_key = str(raw_fields.get("Zotero_Item_Key") or "").strip()
    filename_stem = path.stem
    if zotero_key:
        article_id = zotero_key
    else:
        key_m = re.search(r"\(([A-Z0-9]{8})\)$", filename_stem)
        article_id = key_m.group(1) if key_m else filename_stem

    # ── Year ───────────────────────────────────────────────────────────────
    year_raw = str(raw_fields.get("Year") or "").strip()
    if year_raw and year_raw.lower() not in ("n.d.", "nd", "none", "null", ""):
        try:
            year = int(year_raw)
            year_status = "known"
            decade = f"{(year // 10) * 10}s"
        except ValueError:
            year = None
            year_status = "not_dated"
            decade = "n.d."
            warnings.append(f"Unparseable year: {year_raw!r}")
    else:
        year = None
        year_status = "not_dated"
        decade = "n.d."

    # ── Controlled metadata fields ─────────────────────────────────────────
    controlled: dict[str, str | list[str] | None] = {}
    for field in schema_fields:
        val = raw_fields.get(field)
        if isinstance(val, list):
            # Deduplicate, preserve order, strip whitespace
            seen: set[str] = set()
            clean: list[str] = []
            for v in val:
                v = v.strip()
                if v and v not in seen:
                    seen.add(v)
                    clean.append(v)
            controlled[field] = clean if clean else None
        elif val is not None:
            controlled[field] = str(val).strip() or None
        else:
            controlled[field] = None

    # ── Summary metrics ────────────────────────────────────────────────────
    # Strip the repeated citation heading that appears in many summaries
    summary_body = re.sub(r"^#{1,6}\s+\*\*.*?\*\*\s*\n", "", summary_text, count=1).strip()
    words = summary_body.split() if summary_body else []

    # ── Target sections from contributions ────────────────────────────────
    target_sections = list(
        dict.fromkeys(c["target_section"] for c in contributions if c.get("target_section"))
    )

    return {
        "article_id": article_id,
        "filename": path.name,
        "citation": citation,
        "year": year,
        "year_status": year_status,
        "decade": decade,
        "citation_key": str(raw_fields.get("Citation_Key") or ""),
        "zotero_item_key": zotero_key,
        "better_bibtex_citation_key": str(raw_fields.get("Better_BibTeX_Citation_Key") or ""),
        "attachment_key": str(raw_fields.get("Attachment_Key") or ""),
        "controlled": controlled,
        "summary_text": summary_text,
        "summary_word_count": len(words),
        "summary_character_count": len(summary_body),
        "contributions": contributions,
        "contribution_count": len(contributions),
        "target_sections": target_sections,
        "source_path": str(path),
        "_warnings": warnings,
    }


# ── Field helpers ───────────────────────────────────────────────────────────

def field_key(field_name: str) -> str:
    """Schema field name -> lowercase snake_case key used in CSVs."""
    return re.sub(r"[^a-z0-9]+", "_", field_name.lower()).strip("_")


def value_slug(value: str) -> str:
    """Feature value -> safe column slug."""
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def col_name(fkey: str, vslug: str) -> str:
    return f"{fkey}__{vslug}"


def get_values(record: dict, field_name: str) -> list[str]:
    """Return list of values for a field in a parsed record's controlled dict."""
    val = record["controlled"].get(field_name)
    if val is None:
        return []
    if isinstance(val, list):
        return [v for v in val if v]
    return [str(val)] if str(val).strip() else []


def detect_multi_label_fields(records: list[dict], schema_fields: list[str]) -> set[str]:
    """Return the set of field names where any article has a list value."""
    multi = set()
    for rec in records:
        for field in schema_fields:
            if isinstance(rec["controlled"].get(field), list):
                multi.add(field)
    return multi


# ── Output builders ─────────────────────────────────────────────────────────

def _coerce_year_col(df: pd.DataFrame) -> pd.DataFrame:
    """Cast the year column to nullable Int64 so it displays as integer (or <NA>)."""
    if "year" in df.columns:
        df["year"] = pd.array(df["year"], dtype=pd.Int64Dtype())
    return df


def build_articles_df(
    records: list[dict],
    schema_fields: list[str],
    multi_label_fields: set[str],
    include_summary: bool = True,
) -> pd.DataFrame:
    """Build the main articles DataFrame."""
    rows = []
    for rec in records:
        row: dict = {
            "article_id": rec["article_id"],
            "filename": rec["filename"],
            "citation": rec["citation"],
            "year": rec["year"],
            "year_status": rec["year_status"],
            "decade": rec["decade"],
            "summary_word_count": rec["summary_word_count"],
            "summary_character_count": rec["summary_character_count"],
            "contribution_count": rec["contribution_count"],
            "target_sections": "|".join(rec["target_sections"]) if rec["target_sections"] else "none",
        }
        # Controlled fields (use original schema capitalization)
        for field in schema_fields:
            vals = get_values(rec, field)
            if field in multi_label_fields:
                row[field] = "|".join(vals) if vals else ""
            else:
                row[field] = vals[0] if vals else ""
        row["source_path"] = rec["source_path"]
        if include_summary:
            row["summary_text"] = rec["summary_text"]
        rows.append(row)

    cols = [
        "article_id", "filename", "citation", "year", "year_status", "decade",
        "summary_word_count", "summary_character_count", "contribution_count",
        "target_sections",
    ] + list(schema_fields) + ["source_path"]
    if include_summary:
        cols.append("summary_text")

    return _coerce_year_col(pd.DataFrame(rows, columns=cols))


def build_features_long(
    records: list[dict],
    schema_fields: list[str],
) -> pd.DataFrame:
    """One row per article-feature assignment."""
    rows = []
    for rec in records:
        for field in schema_fields:
            fkey = field_key(field)
            vals = get_values(rec, field)
            for v in vals:
                vslug = value_slug(v)
                rows.append(
                    {
                        "article_id": rec["article_id"],
                        "year": rec["year"],
                        "feature_group": fkey,
                        "feature_value": v,
                        "feature_column": col_name(fkey, vslug),
                        "present": 1,
                    }
                )
    return _coerce_year_col(pd.DataFrame(rows, columns=["article_id", "year", "feature_group", "feature_value", "feature_column", "present"]))


def build_feature_matrix(
    records: list[dict],
    schema_fields: list[str],
    feature_cols: list[str],  # ordered list of 'fkey__vslug' column names
) -> pd.DataFrame:
    """Binary article × feature matrix."""
    col_set = set(feature_cols)

    # Build a lookup: (article_id, col_name) -> 1
    assignments: dict[str, set[str]] = defaultdict(set)
    for rec in records:
        for field in schema_fields:
            fkey = field_key(field)
            for v in get_values(rec, field):
                assignments[rec["article_id"]].add(col_name(fkey, value_slug(v)))

    rows = []
    for rec in records:
        row = {
            "article_id": rec["article_id"],
            "year": rec["year"],
            "decade": rec["decade"],
            "citation": rec["citation"],
            "summary_word_count": rec["summary_word_count"],
        }
        for c in feature_cols:
            row[c] = 1 if c in assignments[rec["article_id"]] else 0
        rows.append(row)

    return _coerce_year_col(pd.DataFrame(rows, columns=IDENTITY_COLS + feature_cols))


def build_field_matrix(
    records: list[dict],
    field: str,
    feature_cols: list[str],
) -> pd.DataFrame:
    """Binary matrix for a single schema field."""
    fkey = field_key(field)
    assignments: dict[str, set[str]] = defaultdict(set)
    for rec in records:
        for v in get_values(rec, field):
            assignments[rec["article_id"]].add(col_name(fkey, value_slug(v)))

    rows = []
    for rec in records:
        row = {
            "article_id": rec["article_id"],
            "year": rec["year"],
            "decade": rec["decade"],
            "citation": rec["citation"],
            "summary_word_count": rec["summary_word_count"],
        }
        for c in feature_cols:
            row[c] = 1 if c in assignments[rec["article_id"]] else 0
        rows.append(row)

    return _coerce_year_col(pd.DataFrame(rows, columns=IDENTITY_COLS + feature_cols))


def build_feature_counts(
    records: list[dict],
    schema_fields: list[str],
) -> pd.DataFrame:
    """One row per unique (feature_group, feature_value) with counts."""
    total = len(records)
    counts: dict[tuple[str, str], int] = defaultdict(int)
    for rec in records:
        for field in schema_fields:
            fkey = field_key(field)
            for v in get_values(rec, field):
                counts[(fkey, v)] += 1

    rows = []
    for (fkey, v), count in counts.items():
        rows.append(
            {
                "feature_group": fkey,
                "feature_value": v,
                "article_count": count,
                "total_articles": total,
                "article_pct": round(count / total * 100, 1) if total else 0,
            }
        )

    df = pd.DataFrame(rows, columns=["feature_group", "feature_value", "article_count", "total_articles", "article_pct"])

    # Dense descending rank within each feature group
    df["rank_within_group"] = (
        df.groupby("feature_group")["article_count"]
        .rank(method="dense", ascending=False)
        .astype(int)
    )

    return df.sort_values(["feature_group", "rank_within_group"]).reset_index(drop=True)


def build_publication_year_counts(records: list[dict]) -> pd.DataFrame:
    total = len(records)
    year_counts: dict = defaultdict(int)
    for rec in records:
        label = str(rec["year"]) if rec["year"] is not None else "n.d."
        year_counts[label] += 1

    rows = []
    for label, count in sorted(
        year_counts.items(),
        key=lambda x: (x[0] == "n.d.", int(x[0]) if x[0] != "n.d." else 9999),
    ):
        rows.append(
            {
                "year_label": label,
                "article_count": count,
                "article_pct": round(count / total * 100, 1) if total else 0,
            }
        )

    return pd.DataFrame(rows, columns=["year_label", "article_count", "article_pct"])


def build_feature_year_counts(
    records: list[dict],
    schema_fields: list[str],
) -> pd.DataFrame:
    """One row per year × feature_group × feature_value combination."""
    year_totals: dict[int | None, int] = defaultdict(int)
    for rec in records:
        year_totals[rec["year"]] += 1

    counts: dict[tuple, int] = defaultdict(int)
    for rec in records:
        for field in schema_fields:
            fkey = field_key(field)
            for v in get_values(rec, field):
                counts[(rec["year"], fkey, v)] += 1

    rows = []
    for (year, fkey, v), count in counts.items():
        denom = year_totals[year]
        rows.append(
            {
                "year": year,
                "feature_group": fkey,
                "feature_value": v,
                "article_count": count,
                "total_articles": denom,
                "article_pct_in_year": round(count / denom * 100, 1) if denom else 0,
            }
        )

    return _coerce_year_col(pd.DataFrame(
        rows,
        columns=["year", "feature_group", "feature_value", "article_count", "total_articles", "article_pct_in_year"],
    ).sort_values(["year", "feature_group", "article_count"], ascending=[True, True, False]).reset_index(drop=True))


def build_feature_cooccurrence(
    records: list[dict],
    schema_fields: list[str],
) -> pd.DataFrame:
    """Unordered pairs of features co-occurring in the same article, ≥2 articles."""
    total = len(records)
    pair_counts: dict[tuple, int] = defaultdict(int)

    for rec in records:
        assignments = []
        for field in schema_fields:
            fkey = field_key(field)
            for v in get_values(rec, field):
                assignments.append((fkey, v))

        for (fa, va), (fb, vb) in combinations(assignments, 2):
            # Canonical ordering
            pair = tuple(sorted([(fa, va), (fb, vb)]))
            pair_counts[pair] += 1

    rows = []
    for ((fa, va), (fb, vb)), count in pair_counts.items():
        if count >= 2:
            rows.append(
                {
                    "feature_group_a": fa,
                    "feature_value_a": va,
                    "feature_group_b": fb,
                    "feature_value_b": vb,
                    "article_count": count,
                    "article_pct": round(count / total * 100, 1) if total else 0,
                }
            )

    df = pd.DataFrame(
        rows,
        columns=["feature_group_a", "feature_value_a", "feature_group_b", "feature_value_b", "article_count", "article_pct"],
    )
    return df.sort_values("article_count", ascending=False).reset_index(drop=True)


def build_contributions_df(records: list[dict]) -> pd.DataFrame:
    rows = []
    for rec in records:
        for i, contrib in enumerate(rec["contributions"], start=1):
            rows.append(
                {
                    "article_id": rec["article_id"],
                    "contribution_number": i,
                    "target_section": contrib.get("target_section", ""),
                    "contribution_text": contrib.get("contribution_text", ""),
                }
            )
    return pd.DataFrame(
        rows, columns=["article_id", "contribution_number", "target_section", "contribution_text"]
    )


def build_dataset_summary(records: list[dict], features_long: pd.DataFrame) -> dict:
    known_years = [r["year"] for r in records if r["year"] is not None]
    word_counts = [r["summary_word_count"] for r in records]
    articles_with_contribs = sum(1 for r in records if r["contribution_count"] > 0)
    not_dated = sum(1 for r in records if r["year"] is None)

    return {
        "total_articles": len(records),
        "earliest_year": min(known_years) if known_years else None,
        "latest_year": max(known_years) if known_years else None,
        "unique_feature_labels": features_long["feature_value"].nunique(),
        "total_article_feature_assignments": len(features_long),
        "mean_features_per_article": round(
            len(features_long) / len(records), 3
        ) if records else 0,
        "median_summary_word_count": float(pd.Series(word_counts).median()) if word_counts else 0,
        "articles_with_contributions": articles_with_contribs,
        "not_dated_articles": not_dated,
    }


# ── Excel workbook ──────────────────────────────────────────────────────────

def write_excel(
    records: list[dict],
    schema_fields: list[str],
    field_values: dict[str, list[str]],
    articles_core: pd.DataFrame,
    feature_counts: pd.DataFrame,
    pub_year_counts: pd.DataFrame,
    dataset_summary: dict,
    data_dict: pd.DataFrame,
    output_path: Path,
) -> None:
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [skip] openpyxl not installed – skipping Excel workbook")
        return

    wb = openpyxl.Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")

    def _write_df(ws, df: pd.DataFrame) -> None:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(wrap_text=True)
        ws.freeze_panes = ws.cell(row=2, column=1)
        ws.auto_filter.ref = ws.dimensions

    def _autofit(ws) -> None:
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0) for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    # Dashboard sheet
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.append(["GBLS Corpus Metrics – Dashboard"])
    ws_dash["A1"].font = Font(bold=True, size=14)
    ws_dash.append([])
    ws_dash.append(["Metric", "Value"])
    for k, v in dataset_summary.items():
        ws_dash.append([k.replace("_", " ").title(), v])
    ws_dash.append([])
    ws_dash.append(["Feature Group", "Unique Values", "Total Assignments"])
    for field in schema_fields:
        fkey = field_key(field)
        grp = feature_counts[feature_counts["feature_group"] == fkey]
        ws_dash.append([field, len(grp), int(grp["article_count"].sum())])
    _autofit(ws_dash)

    # Year Counts
    ws_yr = wb.create_sheet("Year Counts")
    _write_df(ws_yr, pub_year_counts)
    _autofit(ws_yr)

    # All Feature Counts
    ws_fc = wb.create_sheet("All Feature Counts")
    _write_df(ws_fc, feature_counts)
    _autofit(ws_fc)

    # Data Dictionary
    ws_dd = wb.create_sheet("Data Dictionary")
    _write_df(ws_dd, data_dict)
    _autofit(ws_dd)

    # Per-field sheets
    for field in schema_fields:
        fkey = field_key(field)
        sheet_name = field[:31]  # Excel max 31 chars
        ws_f = wb.create_sheet(sheet_name)
        grp = feature_counts[feature_counts["feature_group"] == fkey].drop(
            columns=["feature_group"]
        )
        _write_df(ws_f, grp)
        _autofit(ws_f)

    wb.save(output_path)


# ── JS explorer bundle ──────────────────────────────────────────────────────

def write_js_bundle(
    records: list[dict],
    schema_fields: list[str],
    multi_label_fields: set[str],
    dataset_summary: dict,
    feature_counts: pd.DataFrame,
    pub_year_counts: pd.DataFrame,
    feature_year_counts: pd.DataFrame,
    output_path: Path,
) -> None:
    def _label(field: str) -> str:
        return field.replace("_", " ").title()

    years_known = sorted({r["year"] for r in records if r["year"] is not None})

    feature_groups = []
    for field in schema_fields:
        fkey = field_key(field)
        feature_groups.append(
            {
                "key": fkey,
                "label": _label(field),
                "column": fkey,
                "multi_value": field in multi_label_fields,
            }
        )

    articles_out = []
    for rec in records:
        art: dict = {
            "article_id": rec["article_id"],
            "year": rec["year"],
            "citation": rec["citation"],
        }
        for field in schema_fields:
            fkey = field_key(field)
            vals = get_values(rec, field)
            art[fkey] = "|".join(vals) if field in multi_label_fields else (vals[0] if vals else "")
        articles_out.append(art)

    fc_list = [
        {
            "feature_group": row["feature_group"],
            "feature_value": row["feature_value"],
            "article_count": row["article_count"],
            "article_pct": row["article_pct"],
        }
        for _, row in feature_counts.iterrows()
    ]

    yr_list = [
        {"year_label": row["year_label"], "article_count": row["article_count"]}
        for _, row in pub_year_counts.iterrows()
    ]

    fyr_list = [
        {
            "year": row["year"],
            "feature_group": row["feature_group"],
            "feature_value": row["feature_value"],
            "article_count": row["article_count"],
            "article_pct_in_year": row["article_pct_in_year"],
        }
        for _, row in feature_year_counts.iterrows()
    ]

    payload = {
        "summary": dataset_summary,
        "years": years_known,
        "featureGroups": feature_groups,
        "articles": articles_out,
        "featureCounts": fc_list,
        "publicationYearCounts": yr_list,
        "featureYearCounts": fyr_list,
    }

    js = "window.GBLS_METRICS = " + json.dumps(payload, ensure_ascii=False, default=str) + ";\n"
    output_path.write_text(js, encoding="utf-8")


# ── Validation ──────────────────────────────────────────────────────────────

def validate(
    source_files: list[Path],
    records: list[dict],
    parse_warnings: list[str],
    articles_df: pd.DataFrame,
    features_long: pd.DataFrame,
    feature_counts: pd.DataFrame,
    schema_fields: list[str],
) -> dict:
    total_src = len(source_files)
    total_parsed = len(records)
    article_ids = [r["article_id"] for r in records]
    unique_ids = len(set(article_ids))
    duplicate_ids = total_parsed - unique_ids
    missing_year = sum(1 for r in records if r["year"] is None)

    issues = list(parse_warnings)

    if total_parsed != total_src:
        issues.append(f"Source count ({total_src}) != parsed count ({total_parsed})")

    blank_ids = [i for i in article_ids if not i or not str(i).strip()]
    if blank_ids:
        issues.append(f"{len(blank_ids)} blank article_id values")

    if duplicate_ids:
        from collections import Counter
        dupes = [k for k, v in Counter(article_ids).items() if v > 1]
        issues.append(f"Duplicate article_ids: {dupes}")

    if features_long["present"].ne(1).any():
        issues.append("features_long.present contains values other than 1")

    matrix_rows = len(articles_df)
    if matrix_rows != total_parsed:
        issues.append(f"Matrix rows ({matrix_rows}) != article count ({total_parsed})")

    long_sum = int(features_long.groupby("feature_group")["present"].sum().sum())
    counts_sum = int(feature_counts["article_count"].sum())
    if long_sum != counts_sum:
        issues.append(
            f"Sum of feature_counts.article_count ({counts_sum}) != "
            f"long-table assignments ({long_sum})"
        )

    # Validate all CSVs loadable by pandas
    csv_load_errors = []
    for f in METRICS_DIR.glob("*.csv"):
        try:
            pd.read_csv(f)
        except Exception as exc:
            csv_load_errors.append(f"{f.name}: {exc}")
    if csv_load_errors:
        issues.extend(csv_load_errors)

    return {
        "source_markdown_count": total_src,
        "parsed_article_count": total_parsed,
        "unique_article_id_count": unique_ids,
        "duplicate_article_id_count": duplicate_ids,
        "missing_year_count": missing_year,
        "article_feature_assignment_count": len(features_long),
        "feature_counts_sum": counts_sum,
        "matrix_row_count": matrix_rows,
        "binary_feature_column_count": len(features_long["feature_column"].unique()),
        "parse_warnings": issues,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


# ── Data dictionary builder ─────────────────────────────────────────────────

def build_data_dictionary(schema_fields: list[str], generated_files: list[str]) -> pd.DataFrame:
    rows = [
        ("articles.csv", "Full article table with summary text", "article_id"),
        ("articles_core.csv", "Article table without summary text", "article_id"),
        ("article_features_long.csv", "One row per article-feature assignment (present=1)", "article_id + feature_column"),
        ("article_feature_matrix.csv", "Binary article × feature matrix across all feature groups", "article_id"),
        ("contributions.csv", "One row per extracted proposed contribution", "article_id + contribution_number"),
        ("feature_counts.csv", "Aggregate counts per feature_group + feature_value", "feature_group + feature_value"),
        ("publication_year_counts.csv", "Article counts by publication year", "year_label"),
        ("feature_year_counts.csv", "Feature counts broken down by publication year", "year + feature_group + feature_value"),
        ("feature_cooccurrence.csv", "Unordered pairs of features co-occurring in ≥2 articles", "feature pair"),
        ("dataset_summary.csv", "Headline corpus metrics", "metric"),
        ("dataset_summary.json", "Same as dataset_summary.csv in JSON", "n/a"),
        ("data_dictionary.csv", "This file – documents all generated artifacts", "file"),
        ("feature_datasets_readme.md", "Human-readable guide to the metric files", "n/a"),
        ("validation_report.json", "Parse and consistency validation results", "n/a"),
        ("gbls_feature_overview.xlsx", "Excel workbook with dashboard and per-field sheets", "n/a"),
        ("generated_artifact_manifest.json", "Manifest of all generated files with sizes and timestamp", "n/a"),
    ]
    for field in schema_fields:
        fkey = field_key(field)
        rows.append((f"{fkey}_matrix.csv", f"Binary article × {field} matrix", "article_id"))
        rows.append((f"{fkey}_counts.csv", f"Aggregate counts for {field}", "feature_group + feature_value"))

    return pd.DataFrame(rows, columns=["file", "description", "primary_key"])


# ── Artifact manifest ───────────────────────────────────────────────────────

def write_manifest(generated_files: list[Path], schema_ref: str) -> None:
    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "schema_reference": schema_ref,
        "files_generated": {
            f.name: f.stat().st_size for f in sorted(generated_files) if f.exists()
        },
    }
    manifest_path = METRICS_DIR / "generated_artifact_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")


# ── Obsolete file cleanup ───────────────────────────────────────────────────

def cleanup_obsolete_files(current_files: set[str]) -> list[str]:
    """
    Remove files in METRICS_DIR that match schema-derived naming patterns
    but are NOT in the current generated set (stale from old schema runs).
    Returns list of removed filenames.
    """
    removed = []
    stale_patterns = [
        re.compile(r"^[a-z_]+_matrix\.csv$"),
        re.compile(r"^[a-z_]+_counts\.csv$"),
        re.compile(r"^[a-z_]+_x_[a-z_]+_matrix\.csv$"),  # cross-matrices from old runs
    ]
    for f in METRICS_DIR.glob("*.csv"):
        if f.name in current_files:
            continue
        if any(p.match(f.name) for p in stale_patterns):
            f.unlink()
            removed.append(f.name)
    return removed


# ── Main ────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=" * 60)
    print("GBLS Corpus Metrics Builder")
    print("=" * 60)

    # ── Ensure output directory exists ─────────────────────────────────────
    METRICS_DIR.mkdir(parents=True, exist_ok=True)

    # ── Parse schema ───────────────────────────────────────────────────────
    print("\n[1/9] Parsing schema...")
    schema_fields, field_values = parse_schema(HUMAN_SOURCES_DIR)
    schema_file = find_schema_file(HUMAN_SOURCES_DIR)
    schema_ref = str(schema_file.relative_to(PROJECT_ROOT))
    print(f"  Schema file : {schema_file.name}")
    print(f"  Fields      : {', '.join(schema_fields)}")

    # ── Discover source files ──────────────────────────────────────────────
    print("\n[2/9] Discovering source files...")
    all_md = sorted(SUMMARIES_DIR.glob("*.md"))
    source_files = [
        f for f in all_md
        if not f.name.startswith(".")
        and f.name.lower() != "template.md"
    ]
    print(f"  Found {len(source_files)} source files (excluded template.md and hidden files)")

    # ── Parse each source file ─────────────────────────────────────────────
    print("\n[3/9] Parsing source files...")
    records: list[dict] = []
    parse_warnings: list[str] = []
    errors: list[str] = []

    for path in source_files:
        rec = parse_summary_file(path, schema_fields)
        if "_error" in rec:
            errors.append(f"{path.name}: {rec['_error']}")
            print(f"  ERROR: {path.name}: {rec['_error']}", file=sys.stderr)
            continue
        if rec["_warnings"]:
            for w in rec["_warnings"]:
                parse_warnings.append(f"{path.name}: {w}")
        records.append(rec)

    if errors:
        print(f"\n  {len(errors)} files failed to parse – stopping.")
        for e in errors:
            print(f"    {e}", file=sys.stderr)
        sys.exit(1)

    print(f"  Parsed {len(records)} articles ({len(parse_warnings)} warnings)")

    # ── Detect multi-label fields ──────────────────────────────────────────
    multi_label_fields = detect_multi_label_fields(records, schema_fields)
    print(f"  Multi-label fields: {', '.join(sorted(multi_label_fields)) or '(none)'}")

    # ── Build core data structures ─────────────────────────────────────────
    print("\n[4/9] Building data structures...")
    features_long = build_features_long(records, schema_fields)
    feature_cols_ordered = sorted(features_long["feature_column"].unique())

    articles_df = build_articles_df(records, schema_fields, multi_label_fields, include_summary=True)
    articles_core_df = build_articles_df(records, schema_fields, multi_label_fields, include_summary=False)
    feature_matrix_df = build_feature_matrix(records, schema_fields, feature_cols_ordered)
    feature_counts_df = build_feature_counts(records, schema_fields)
    pub_year_df = build_publication_year_counts(records)
    feature_year_df = build_feature_year_counts(records, schema_fields)
    cooccurrence_df = build_feature_cooccurrence(records, schema_fields)
    contributions_df = build_contributions_df(records)
    dataset_summary = build_dataset_summary(records, features_long)

    print(f"  Articles          : {len(records)}")
    print(f"  Feature assignments: {len(features_long)}")
    print(f"  Feature columns   : {len(feature_cols_ordered)}")

    # ── Write CSV outputs ──────────────────────────────────────────────────
    print("\n[5/9] Writing CSV files...")
    generated_files: list[Path] = []

    def save_csv(df: pd.DataFrame, name: str) -> Path:
        path = METRICS_DIR / name
        df.to_csv(path, index=False, encoding="utf-8")
        generated_files.append(path)
        return path

    save_csv(articles_df, "articles.csv")
    save_csv(articles_core_df, "articles_core.csv")
    save_csv(features_long, "article_features_long.csv")
    save_csv(feature_matrix_df, "article_feature_matrix.csv")
    save_csv(contributions_df, "contributions.csv")
    save_csv(feature_counts_df, "feature_counts.csv")
    save_csv(pub_year_df, "publication_year_counts.csv")
    save_csv(feature_year_df, "feature_year_counts.csv")
    save_csv(cooccurrence_df, "feature_cooccurrence.csv")

    # Per-field matrices and count tables
    for field in schema_fields:
        fkey = field_key(field)
        field_feature_cols = [c for c in feature_cols_ordered if c.startswith(fkey + "__")]
        field_matrix = build_field_matrix(records, field, field_feature_cols)
        save_csv(field_matrix, f"{fkey}_matrix.csv")
        field_counts = feature_counts_df[feature_counts_df["feature_group"] == fkey].copy()
        save_csv(field_counts, f"{fkey}_counts.csv")

    # dataset_summary.csv
    ds_df = pd.DataFrame(
        [{"metric": k, "value": v} for k, v in dataset_summary.items()]
    )
    save_csv(ds_df, "dataset_summary.csv")

    # dataset_summary.json
    ds_json_path = METRICS_DIR / "dataset_summary.json"
    ds_json_path.write_text(json.dumps(dataset_summary, indent=2, default=str), encoding="utf-8")
    generated_files.append(ds_json_path)

    print(f"  Wrote {len(generated_files)} CSV/JSON files")

    # ── Data dictionary ────────────────────────────────────────────────────
    print("\n[6/9] Writing supporting files...")
    data_dict_df = build_data_dictionary(schema_fields, [f.name for f in generated_files])
    save_csv(data_dict_df, "data_dictionary.csv")

    # README
    readme_path = METRICS_DIR / "feature_datasets_readme.md"
    readme_text = f"""# GBLS Feature Datasets

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}
Articles: {len(records)} | Feature assignments: {len(features_long)}

## Quick start (Python / Colab)

```python
import pandas as pd

metrics_dir = "2_calculated_metrics/gbls_corpus_metrics"
articles = pd.read_csv(f"{{metrics_dir}}/articles_core.csv")
features = pd.read_csv(f"{{metrics_dir}}/article_features_long.csv")
matrix   = pd.read_csv(f"{{metrics_dir}}/article_feature_matrix.csv")
```

## Key files

| File | Grain | Purpose |
|------|-------|---------|
| `articles_core.csv` | 1 row/article | All metadata fields; pipe-delimited for multi-label |
| `articles.csv` | 1 row/article | Same + full summary text |
| `article_features_long.csv` | 1 row/assignment | Tidy format; filter by `feature_group` |
| `article_feature_matrix.csv` | 1 row/article | Binary columns across all feature groups |
| `feature_counts.csv` | 1 row/value | Article counts and percentages per feature |
| `feature_cooccurrence.csv` | 1 row/pair | Co-occurring feature pairs (≥2 articles) |

## Counting conventions

- **Counts represent coded article presence**, not prose term frequency.
  One article contributes at most 1 count per feature value.
- **Multi-label fields** ({', '.join(sorted(multi_label_fields))})
  may contribute multiple assignments per article — one per coded value.
- `article_pct` = article_count / total_articles × 100
- `article_pct_in_year` uses same-year article count as denominator.

## Schema fields

{chr(10).join(f'- **{f}**' + (' (multi-label)' if f in multi_label_fields else '') for f in schema_fields)}

## Undated records

{sum(1 for r in records if r['year'] is None)} article(s) have no year and are labelled `n.d.` in year summaries.
"""
    readme_path.write_text(readme_text, encoding="utf-8")
    generated_files.append(readme_path)

    # ── Validation report ──────────────────────────────────────────────────
    print("\n[7/9] Validating...")
    report = validate(
        source_files, records, parse_warnings,
        articles_core_df, features_long, feature_counts_df, schema_fields
    )
    vr_path = METRICS_DIR / "validation_report.json"
    vr_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    generated_files.append(vr_path)

    issue_count = len(report["parse_warnings"])
    if issue_count:
        print(f"  {issue_count} warnings/issues:")
        for w in report["parse_warnings"][:10]:
            print(f"    {w}")
        if issue_count > 10:
            print(f"    ... and {issue_count - 10} more (see validation_report.json)")
    else:
        print("  All checks passed.")

    # ── Excel workbook ─────────────────────────────────────────────────────
    print("\n[8/9] Writing Excel workbook...")
    xlsx_path = METRICS_DIR / "gbls_feature_overview.xlsx"
    write_excel(
        records, schema_fields, field_values,
        articles_core_df, feature_counts_df, pub_year_df,
        dataset_summary, data_dict_df, xlsx_path
    )
    if xlsx_path.exists():
        generated_files.append(xlsx_path)
        print(f"  Wrote {xlsx_path.name}")

    # ── JS explorer bundle ─────────────────────────────────────────────────
    print("\n[9/9] Refreshing JS explorer bundle...")
    if EXPLORER_HTML.exists() and EXPLORER_JS.exists():
        write_js_bundle(
            records, schema_fields, multi_label_fields,
            dataset_summary, feature_counts_df, pub_year_df,
            feature_year_df, EXPLORER_DATA
        )
        print(f"  Wrote {EXPLORER_DATA.name}")
        generated_files.append(EXPLORER_DATA)
        if EXPLORER_DATA_CODER.exists():
            import shutil as _shutil
            _shutil.copy2(EXPLORER_DATA, EXPLORER_DATA_CODER)
            print(f"  Copied → {EXPLORER_DATA_CODER.relative_to(PROJECT_ROOT)}")
    else:
        print("  Explorer HTML/JS not found – skipping bundle.")

    # ── Cleanup obsolete schema-derived files ──────────────────────────────
    current_names = {f.name for f in generated_files}
    removed = cleanup_obsolete_files(current_names)
    if removed:
        print(f"\n  Removed {len(removed)} obsolete files: {', '.join(removed)}")

    # ── Manifest ───────────────────────────────────────────────────────────
    manifest_path = METRICS_DIR / "generated_artifact_manifest.json"
    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "schema_reference": schema_ref,
        "total_articles": len(records),
        "files_generated": {
            f.name: f.stat().st_size for f in sorted(generated_files) if f.exists()
        },
    }
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    # ── Summary ────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("BUILD COMPLETE")
    print("=" * 60)
    print(f"  Source summaries parsed : {len(records)}")
    print(f"  Unique articles         : {len(set(r['article_id'] for r in records))}")
    print(f"  Feature assignments     : {len(features_long)}")
    print(f"  Unique coded labels     : {features_long['feature_value'].nunique()}")
    print(f"  Undated sources         : {sum(1 for r in records if r['year'] is None)}")
    print(f"  Workbook created        : {xlsx_path.exists()}")
    js_refreshed = EXPLORER_HTML.exists() and EXPLORER_JS.exists()
    print(f"  JS explorer refreshed   : {js_refreshed}")
    print(f"  Output folder           : {METRICS_DIR}")
    print(f"  Files generated         : {len(generated_files)}")
    print()


if __name__ == "__main__":
    main()
